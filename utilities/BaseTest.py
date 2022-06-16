import inspect
import logging

import openpyxl
import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@pytest.mark.usefixtures("setup")
class BaseTest:

    testDataPath = "C:\\Users\\T951881\\PycharmProjects\\PythonSeleniumFramework\\TestData\\TestData.xlsx"

    def getLogger(self):
        # __name__ parameter is responsible for printing log with testcase name
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        filehandler = logging.FileHandler(
            'C:\\Users\\T951881\\PycharmProjects\\PythonSeleniumFramework\\reports_n_logs\\logFile.log')  # File path is specified
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s : %(message)s")
        filehandler.setFormatter(formatter)  # formatter is attached to the file itself

        logger.addHandler(filehandler)  # Filehandler object is passed

        # log order is  debug-->info-->warning-->error-->critical
        logger.setLevel(logging.DEBUG)
        # it will skip logging debug messages and just log info->warning->error->critical
        return logger

    def refreshBrowser(self):
        self.driver.refresh()

    def verifyElementPresence(self, text):
        # Explicit Waits
        wait = WebDriverWait(self.driver, 10)
        wait.until(EC.presence_of_element_located((By.LINK_TEXT, text)))

    def selectDDOptionsByText(self, locator, text):
        dropdown = Select(locator)
        dropdown.select_by_visible_text(text)

    @staticmethod
    def getTestDataSetAllRows(testDataPath, sheetName):
        Dict = {}
        dataSetList = []
        wb = openpyxl.load_workbook(testDataPath)
        ws = wb[sheetName]

        for i in range(1, ws.max_row):
            # if sheet.cell(row=i, column=1).value == "TC1":
            for j in range(1, ws.max_column + 1):
                Dict[ws.cell(row=1, column=j).value] = ws.cell(row=i + 1, column=j).value
            dataSetList.append(Dict)
            Dict = {}
        print(dataSetList)
        return dataSetList

    @staticmethod
    def getTestDataByID(testDataPath, sheetName, test_case_id):
        Dict = {}
        wb = openpyxl.load_workbook(testDataPath)
        ws = wb[sheetName]

        for i in range(1, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == test_case_id:
                for j in range(1, ws.max_column + 1):
                    Dict[ws.cell(row=1, column=j).value] = ws.cell(row=i, column=j).value
        print(Dict)
        return [Dict]
